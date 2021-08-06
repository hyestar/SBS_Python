item1 = {"품목" : "바나나", "가격" : 5000,  "수량" : 1, "날짜" : "2021년 08월 5일"}
item2 = {"품목" : "커피"  , "가격" : 3000,  "수량" : 2, "날짜" : "2021년 08월 5일"}
item3 = {"품목" : "책"    , "가격" : 15000, "수량" : 1, "날짜" : "2021년 08월 6일"}
item4 = {"품목" : "펜"    , "가격" : 2000,  "수량" : 4, "날짜" : "2021년 08월 6일"}

items = [item1, item2, item3, item4]

from openpyxl import *

w1 = Workbook() # class
ws = w1.active

ws["A1"]="품목"
ws["B1"]="가격"
ws["C1"]="수량"
ws["D1"]="총지출"
ws["E1"]="날짜"
i=2
for item in items :
    ws["A"+str(i)] = item["품목"]
    ws["B"+str(i)] = item["가격"]
    ws["C"+str(i)] = item["수량"]
    ws["D"+str(i)] = item["가격"]*item["수량"]   
    ws["E"+str(i)] = item["날짜"]
    i+=1 
w1.save("item.xlsx")

w2 = load_workbook("item.xlsx")  
ws2= w2.active
# 영역으로 꺼내오기
itemList=[]
# 행은 데이터가 추가 될수록 변경되니깐 직접 만들어야 함
rowNum = ws.max_row  # 마지막 행의 번호
for row in ws["A2":"E"+str(rowNum)] :
    item = {"품목":row[0].value,
    "가격":row[1].value,
    "수량":row[2].value,
    "가격":row[3].value,
    "날짜":row[4].value
    }
    itemList.append(item)
print(itemList)

w2.save("item.xlsx")