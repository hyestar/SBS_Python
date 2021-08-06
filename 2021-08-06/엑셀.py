# pip install openpyxl 터미널에 입력

# 모듈 import
from openpyxl import *

# 엑셀 추가
w1 = Workbook() # class
ws = w1.active

# 엑셀 저장
w1.save("test.xlsx")

# 엑셀 로드
w2 = load_workbook("test.xlsx")
ws2 = w2.active # 활성화

# 하나의 지정된 셀을 선택해서 편집
ws2["A1"]="hello"

# 리스트, 튜플
list1 = [1,2,3,4]
tuple1 = 1,2,3,4
# 리스트는 변경가능하지만 튜플은 값변경 불가!!


# 많은 셀을 선택해서 편집(영역필요)
num = 1
for row in ws2["A1":"C3"] : # 열만 먼저 꺼내고
    for cell in row :   # 행꺼내기
        cell.value = num
        num+=1

w2.save("test.xlsx")

# 원하는 만큼 셀을 조회
for row in ws2.iter_rows(min_row=1, min_col=2, max_row=3) : # 1행부터 3행까지 2개의 열 가져오기
    for cell in row:
        print(cell)